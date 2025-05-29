import csv
import openpyxl

def read_csv(file_path):
    """Reads a CSV file and returns its content as a list of dictionaries."""
    with open(file_path, mode='r', newline='', encoding='utf-8') as csvfile:
        reader = csv.DictReader(csvfile)
        return [row for row in reader]

def main():
    file_path = 'vn.csv'  # Replace with your CSV file path
    data = read_csv(file_path)
    
    # Create an Excel file
    
    # Create a new workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Servers"
    
    # Add headers
    headers = ["server_id", "server_name", "status", "ipv4", "port"]
    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_num, value=header)
    
    # Add data rows
    for id, row in enumerate(data):
        server_id = "Server " + str(id + 1)
        server_name = row.get('name', '')
        if server_name == '':
            server_name = server_id

        status = 'Off'
        ipv4 = row.get('ip_address', '')
        port = row.get('port', '')
        
        worksheet.cell(row=id+2, column=1, value=server_id)
        worksheet.cell(row=id+2, column=2, value=server_name)
        worksheet.cell(row=id+2, column=3, value=status)
        worksheet.cell(row=id+2, column=4, value=ipv4)
        worksheet.cell(row=id+2, column=5, value=port)
    
    # Save the workbook
    workbook.save('output.xlsx')

if __name__ == "__main__":
    main()